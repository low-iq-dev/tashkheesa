import openpyxl
import json

src = '/Users/ziadelwahsh/tashkheesa-portal/docs/pricing/tashkheesa_pricing_v2.xlsx'
dst = '/Users/ziadelwahsh/tashkheesa-portal/docs/pricing/tashkheesa_pricing_v2.json'

wb = openpyxl.load_workbook(src)
ws = wb['Master Pricing (EGP)']

services = []
unpriced = []

for r in range(4, ws.max_row + 1):
    num = ws.cell(row=r, column=1).value
    specialty = ws.cell(row=r, column=2).value
    name = ws.cell(row=r, column=3).value
    launch = ws.cell(row=r, column=4).value
    tier = ws.cell(row=r, column=5).value
    shifa_cost = ws.cell(row=r, column=6).value
    tashk_price = ws.cell(row=r, column=7).value
    doctor_fee = ws.cell(row=r, column=8).value
    keep = ws.cell(row=r, column=9).value
    status = ws.cell(row=r, column=10).value

    if not name or not specialty:
        continue
    if not isinstance(name, str) or not isinstance(specialty, str):
        continue
    if 'price floor' in str(specialty).lower() or '🔺' in str(specialty):
        continue  # footnote rows

    record = {
        'specialty': specialty.strip(),
        'name': name.strip(),
        'tier': (tier or 'Simple').strip() if isinstance(tier, str) else 'Simple',
        'shifa_cost': shifa_cost if isinstance(shifa_cost, (int, float)) else None,
        'tashkheesa_price': tashk_price if isinstance(tashk_price, (int, float)) else None,
        'doctor_fee': doctor_fee if isinstance(doctor_fee, (int, float)) else None,
        'launch_flag': str(launch) if launch else None,
        'status': str(status) if status else None,
    }

    if record['tashkheesa_price'] is None or record['doctor_fee'] is None:
        unpriced.append(record)
    else:
        services.append(record)

out = {
    'source': src,
    'priced_services': services,
    'unpriced_services': unpriced,
    'priced_count': len(services),
    'unpriced_count': len(unpriced),
}

with open(dst, 'w') as f:
    json.dump(out, f, indent=2, ensure_ascii=False)

print(f'Wrote {dst}')
print(f'  Priced:   {len(services)}')
print(f'  Unpriced: {len(unpriced)}')

# Validate the 20% rule
mismatches = []
for s in services:
    expected = round(s['tashkheesa_price'] * 0.20)
    if abs(s['doctor_fee'] - expected) > 1:
        mismatches.append(s)
if mismatches:
    print(f'  WARNING: {len(mismatches)} priced services do not match 20% rule:')
    for m in mismatches[:5]:
        print(f'    - {m["specialty"]} / {m["name"]}: tp={m["tashkheesa_price"]} fee={m["doctor_fee"]}')
else:
    print(f'  ✓ All {len(services)} priced services match 20% rule')
